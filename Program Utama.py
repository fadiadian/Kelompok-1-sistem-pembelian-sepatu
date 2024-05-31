
from pathlib import Path
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, messagebox
import csv
import os
import pandas as pd
from tkinter import ttk
import tkinter as tk
import time
import random
from PIL import Image
import bismillah2 as bs

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"D:\Praktikum Prokom\Tubes clone\Kelompok-1-sistem-pembelian-sepatu\assets\frame0") #ini kynya udah gak kepake, soalnya udah pakai path absolut

def relative_to_assets(path: str) -> str:
    # Mengembalikan path absolut
    return os.path.join(os.getcwd(),'assets', 'frame0', path)


#=============  V  verifikasi dan proses login  V  =================
def verifikasi_login(username, password):
    # Menggunakan path absolut
    file_path = os.path.join(os.getcwd(), 'databaseuser', 'logindatabase.csv')
    
    try:
        with open(file_path, mode='r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                if row['username'] == username and row['password'] == password:
                    return True
        return False
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return False

def proses_login():
    username = entry_username.get()
    password = entry_pw.get()
    
    if verifikasi_login(username, password):
        messagebox.showinfo("Sukses", "Login berhasil!") #kalo main menu udah jadi masukin di baris setelah ini
        window.destroy()
        menu()
    else:
        messagebox.showerror("Error", "Nama pengguna atau kata sandi salah")

#=============  A  verifikasi dan proses login  A  =================
window = Tk()
def program_awal():
    global window
    global entry_username, entry_pw
    global random_data

    window.geometry("1080x720")
    window.configure(bg="#FFFFFF")


    csv_file_path = os.path.join(os.getcwd(), 'Persediaan1.xlsx')  # Ganti dengan path file CSV Anda
    all_data = read_data_from_csv(csv_file_path)
    random_data = all_data.sample(n=8)

    canvas = Canvas(
        window,
        bg="#FFFFFF",
        height=720,
        width=1080,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        386.0,
        300.0,
        image=entry_image_1
    )
    entry_username = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 30 * -1)
    )
    entry_username.place(
        x=93.0,
        y=269.0,
        width=586.0,
        height=60.0
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets("entry_2.png"))
    entry_bg_2 = canvas.create_image(
        386.0,
        427.0,
        image=entry_image_2
    )
    entry_pw = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 30 * -1), show="*"
    )
    entry_pw.place(
        x=93.0,
        y=396.0,
        width=586.0,
        height=60.0
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1.png"))
    button_login = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=proses_login,
        relief="flat"
    )
    button_login.place(
        x=163.0,
        y=549.0,
        width=753.0,
        height=62.0
    )

    canvas.create_text(
        73.0,
        133.0,
        anchor="nw",
        text="Silakan login",
        fill="#000000",
        font=("Segoe UI Bold", 30 * -1, 'bold')
    )

    canvas.create_text(
        73.0,
        239.0,
        anchor="nw",
        text="Masukkan Username",
        fill="#000000",
        font=("Segoe UI SemiBold", 20 * -1, 'bold')
    )

    canvas.create_text(
        73.0,
        58.0,
        anchor="nw",
        text="Selamat datang!",
        fill="#0057FF",
        font=("Segoe UI Bold", 50 * -1, 'bold')
    )

    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    image_1 = canvas.create_image(
        903.0,
        363.0,
        image=image_image_1
    )

    canvas.create_text(
        73.0,
        366.0,
        anchor="nw",
        text="Masukkan Sandi",
        fill="#000000",
        font=("Segoe UI SemiBold", 20 * -1, 'bold')
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2.png"))
    button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("button_2 clicked"),
        relief="flat"
    )
    button_2.place(
        x=549.0,
        y=458.0,
        width=150.0,
        height=30.0
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    image_2 = canvas.create_image(
        917.0,
        82.0,
        image=image_image_2
    )
    window.resizable(False, False)
    window.mainloop()

# main ini nanti dimasukin ke login 
# assets ..._main.png sudah boleh dihapus

# =============== V Ini halaman utama V =================

def read_data_from_csv(file_path):
    reader = pd.read_excel(file_path)
    reader = pd.DataFrame(reader)
    return reader


def daftar_sepatu_menu(row):
   tampil_daftar = random_data['series'].values[row]
   return tampil_daftar

def menu():
    set_sepatu('', '', 0, 0)
    global diskon
    diskon = 0
    window_2 = Tk()

    window_2.geometry("1080x720")
    window_2.configure(bg = "#01041A")

    canvas = Canvas(
        window_2,
        bg = "#01041A",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )
    csv_file_path = os.path.join(os.getcwd(), 'Persediaan1.xlsx')  # Ganti dengan path file CSV Anda
    all_data = read_data_from_csv(csv_file_path)
    random_data = all_data.sample(n=8)
    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("menu_image_1.png"))
    image_1 = canvas.create_image(
        207.0,
        71.0,
        image=image_image_1
    )

    canvas.create_text(
        85.0,
        38.0, #32
        anchor="nw",
        text="Menu",
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1) #25
    )

    menu_button_image_1 = PhotoImage(
        file=relative_to_assets("menu_button_1.png"))
    menu_button_1 = Button(
        image=menu_button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("menu_button_1 clicked"), window_2.destroy(), kasir()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_1.place(
        x=41.0,
        y=80.0,
        width=25.0,
        height=25.0
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("menu_image_2.png"))
    image_2 = canvas.create_image(
        53.0,
        51.0,
        image=image_image_2
    )

    entry_image_1 = PhotoImage(
        file=relative_to_assets("menu_entry_1.png"))
    entry_bg_1 = canvas.create_image(
        175.5,
        91.5,
        image=entry_image_1
    )
    menu_entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    menu_entry_1.place(
        x=97.5,
        y=79.0,
        width=156.0,
        height=23.0
    )

    image_image_3 = PhotoImage(
        file=relative_to_assets("menu_image_3.png"))
    image_3 = canvas.create_image(
        207.0,
        425.0,
        image=image_image_3
    )

    canvas.create_text(
        53.0,
        164.0,
        anchor="nw",
        text="Stok saat ini",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, "bold")
    )
    canvas.create_text(
        53.0,
        218.0,
        anchor="nw",
        text=str(random_data['merek'].values[0])+ '' + str(random_data['series'].values[0])+ ' ukuran '+ str(random_data['ukuran'].values[0]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        53.0,
        258.0,
        anchor="nw",
        text=str(random_data['merek'].values[1])+ '' + str(random_data['series'].values[1])+ ' ukuran '+ str(random_data['ukuran'].values[1]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        54.0,
        298.0,
        anchor="nw",
        text=str(random_data['merek'].values[2])+ '' + str(random_data['series'].values[2])+ ' ukuran '+ str(random_data['ukuran'].values[2]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        54.0,
        338.0,
        anchor="nw",
        text=str(random_data['merek'].values[3])+ '' + str(random_data['series'].values[3])+ ' ukuran '+ str(random_data['ukuran'].values[3]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        54.0,
        378.0,
        anchor="nw",
        text=str(random_data['merek'].values[4])+ '' + str(random_data['series'].values[4])+ ' ukuran '+ str(random_data['ukuran'].values[4]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        55.0,
        418.0,
        anchor="nw",
        text=str(random_data['merek'].values[5])+ '' + str(random_data['series'].values[5])+ ' ukuran '+ str(random_data['ukuran'].values[5]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        54.0,
        498.0,
        anchor="nw",
        text=str(random_data['merek'].values[6])+ '' + str(random_data['series'].values[6])+ ' ukuran '+ str(random_data['ukuran'].values[6]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        53.0,
        458.0,
        anchor="nw",
        text=str(random_data['merek'].values[7])+ '' + str(random_data['series'].values[7])+ ' ukuran '+ str(random_data['ukuran'].values[7]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        218.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[0]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        258.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[1]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        298.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[2]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        338.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[3]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        378.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[4]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        418.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[5]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        498.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[6]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    canvas.create_text(
        350.0,
        458.0,
        anchor="nw",
        text=str(random_data['jumlah'].values[7]),
        fill="#FFFFFF",
        font=("Segoe UI", 20 * -1)
    )

    image_image_4 = PhotoImage(
        file=relative_to_assets("menu_image_4.png"))
    image_4 = canvas.create_image(
        739.0,
        360.0,
        image=image_image_4
    )

    menu_button_image_2 = PhotoImage(
        file=relative_to_assets("menu_button_2.png"))
    menu_button_2 = Button(
        image=menu_button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("menu_button_2 clicked"), load_and_display_excel_data()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_2.place(
        x=484.0,
        y=69.0,
        width=548.9635009765625,
        height=127.0
    )

    canvas.create_text(
        477.0,
        26.0, #16
        anchor="nw",
        text="Toko",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, "bold")
    )

    menu_button_image_3 = PhotoImage(
        file=relative_to_assets("menu_button_3.png"))
    menu_button_3 = Button(
        image=menu_button_image_3,
        borderwidth=0,
        highlightthickness=0,
        command= lambda: [window_2.destroy(), ud_stok()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_3.place(
        x=675.0,
        y=241.0,
        width=160.0,
        height=241.0
    )

    menu_button_image_4 = PhotoImage(
        file=relative_to_assets("menu_button_4.png"))
    menu_button_4 = Button(
        image=menu_button_image_4,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("menu_button_4 clicked"), window_2.destroy() ,tambah_produk()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_4.place(
        x=867.0,
        y=250.0,
        width=160.0,
        height=232.0
    )

    menu_button_image_5 = PhotoImage(
        file=relative_to_assets("menu_button_5.png"))
    menu_button_5 = Button(
        image=menu_button_image_5,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("menu_button_5 clicked"), window_2.destroy(), kasir()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_5.place(
        x=482.0,
        y=250.0,
        width=160.0,
        height=232.0
    )

    canvas.create_text(
        482.0,
        222.0,
        anchor="nw",
        text="Fitur",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, "bold")
    )

    menu_button_image_6 = PhotoImage(
        file=relative_to_assets("menu_button_6.png"))
    menu_button_6 = Button(
        image=menu_button_image_6,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("menu_button_6 clicked"), window_2.destroy(), tradein()],
        relief="flat",
        bg="#1A1E3E",
        activebackground="#1A1E3E"
    )
    menu_button_6.place(
        x=485.0,
        y=558.0,
        width=544.0,
        height=140.0 #height=127.0
    )

    canvas.create_text(
        483.0,
        508.0,#498
        anchor="nw",
        text="Lainnya",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, "bold")
    )
    window_2.resizable(False, False)
    window_2.mainloop()
#==================== V GUI tambah stock V ===================
def ud_stok():
    window_ud_stok = Tk()

    window_ud_stok.geometry("1080x720")
    window_ud_stok.configure(bg = "#01041A")


    canvas = Canvas(
        window_ud_stok,
        bg = "#01041A",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_ud_stock.png"))
    image_1 = canvas.create_image(
        541.0,
        231.0,
        image=image_image_1
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2_ud_stock.png"))
    image_2 = canvas.create_image(
        541.0,
        417.0,
        image=image_image_2
    )

    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1_ud_stock.png"))
    entry_bg_1 = canvas.create_image(
        517.5,
        277.0,
        image=entry_image_1
    )
    ud_stok_entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_1.place(
        x=438.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        423.0,
        206.0,
        anchor="nw",
        text="Merek",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets("entry_2_ud_stock.png"))
    entry_bg_2 = canvas.create_image(
        194.0,
        277.0,
        image=entry_image_2
    )
    ud_stok_entry_2 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_2.place(
        x=71.0,
        y=250.0,
        width=246.0,
        height=52.0
    )

    entry_image_3 = PhotoImage(
        file=relative_to_assets("entry_3_ud_stock.png"))
    entry_bg_3 = canvas.create_image(
        259.5,
        417.0,
        image=entry_image_3
    )
    ud_stok_entry_3 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_3.place(
        x=202.0,
        y=390.0,
        width=115.0,
        height=52.0
    )

    canvas.create_text(
        56.0,
        206.0,
        anchor="nw",
        text="Masukkan ID",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    canvas.create_text(
        56.0,
        400.0,
        anchor="nw",
        text="Jumlah :",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_4 = PhotoImage(
        file=relative_to_assets("entry_4_ud_stock.png"))
    entry_bg_4 = canvas.create_image(
        770.5,
        417.0,
        image=entry_image_4
    )
    ud_stok_entry_4 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_4.place(
        x=672.0,
        y=390.0,
        width=197.0,
        height=52.0
    )

    canvas.create_text(
        422.0,
        400.0,
        anchor="nw",
        text="Harga saat ini :",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5_ud_stock.png"))
    entry_bg_5 = canvas.create_image(
        770.5,
        277.0,
        image=entry_image_5
    )
    ud_stok_entry_5 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_5.place(
        x=672.0,
        y=250.0,
        width=197.0,
        height=52.0
    )

    canvas.create_text(
        657.0,
        206.0,
        anchor="nw",
        text="Seri",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_6 = PhotoImage(
        file=relative_to_assets("entry_6_ud_stock.png"))
    entry_bg_6 = canvas.create_image(
        975.0,
        277.0,
        image=entry_image_6
    )
    ud_stok_entry_6 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    ud_stok_entry_6.place(
        x=946.0,
        y=250.0,
        width=58.0,
        height=52.0
    )

    canvas.create_text(
        931.0,
        206.0,
        anchor="nw",
        text="Ukuran",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    canvas.create_text(
        56.0,
        32.0,
        anchor="nw",
        text="Update Stok",
        fill="#FFFFFF",
        font=("Segoe UI", 40 * -1, 'bold')
    )

    canvas.create_text(
        56.0,
        150.0,
        anchor="nw",
        text="Dengan ID",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, 'bold')
    )

    canvas.create_text(
        423.0,
        150.0,
        anchor="nw",
        text="Manual",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, 'bold')
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_ud_stock.png"))
    ud_stok_button_1 = Button(
        image=button_image_1, #entry 6 adalah  seri, 
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), submit_edit(ud_stok_entry_2, ud_stok_entry_1, ud_stok_entry_5, ud_stok_entry_6, ud_stok_entry_4, ud_stok_entry_3), window_ud_stok.destroy(),menu()],
        bg="#01041A",
        activebackground="#01041A",
        relief="flat"
    )
    ud_stok_button_1.place(
        x=115.0,
        y=554.0,
        width=360.0,
        height=56.0
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2_ud_stock.png"))
    ud_stok_button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_2 clicked"), window_ud_stok.destroy(), tambah_produk()],
        bg="#01041A",
        activebackground="#01041A",
        relief="flat"
    )
    ud_stok_button_2.place(
        x=608.0,
        y=554.0,
        width=360.0,
        height=56.0
    )

    canvas.create_rectangle(
        376.0,
        149.0,
        377.0,
        304.0,
        fill="#FFFFFF",
        outline="")
    window_ud_stok.resizable(False, False)
    window_ud_stok.mainloop()

#=================== V GUI tambah produk baru V ====================
def tambah_produk():
    window = Tk()

    window.geometry("1080x720")
    window.configure(bg = "#01041A")


    canvas = Canvas(
        window,
        bg = "#01041A",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_produk_baru.png"))
    image_1 = canvas.create_image(
        541.0,
        231.0,
        image=image_image_1
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2_produk_baru.png"))
    image_2 = canvas.create_image(
        541.0,
        417.0,
        image=image_image_2
    )

    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1_produk_baru.png"))
    entry_bg_1 = canvas.create_image(
        454.5,
        277.0,
        image=entry_image_1
    )
    tambah_produk_entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_1.place(
        x=375.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        360.0,
        206.0,
        anchor="nw",
        text="Merek",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets("entry_2_produk_baru.png"))
    entry_bg_2 = canvas.create_image(
        212.5,
        277.0,
        image=entry_image_2
    )
    tambah_produk_entry_2 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_2.place(
        x=133.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        206.0,
        anchor="nw",
        text="Jenis",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_3 = PhotoImage(
        file=relative_to_assets("entry_3_produk_baru.png"))
    entry_bg_3 = canvas.create_image(
        259.5,
        417.0,
        image=entry_image_3
    )
    tambah_produk_entry_3 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_3.place(
        x=202.0,
        y=390.0,
        width=115.0,
        height=52.0
    )

    canvas.create_text(
        56.0,
        400.0,
        anchor="nw",
        text="Jumlah :",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_4 = PhotoImage(
        file=relative_to_assets("entry_4_produk_baru.png"))
    entry_bg_4 = canvas.create_image(
        770.5,
        417.0,
        image=entry_image_4
    )
    tambah_produk_entry_4 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_4.place(
        x=672.0,
        y=390.0,
        width=197.0,
        height=52.0
    )

    canvas.create_text(
        541.0,
        400.0,
        anchor="nw",
        text="Harga :",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5_produk_baru.png"))
    entry_bg_5 = canvas.create_image(
        707.5,
        277.0,
        image=entry_image_5
    )
    tambah_produk_entry_5 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_5.place(
        x=609.0,
        y=250.0,
        width=197.0,
        height=52.0
    )

    canvas.create_text(
        594.0,
        206.0,
        anchor="nw",
        text="Seri",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    entry_image_6 = PhotoImage(
        file=relative_to_assets("entry_6_produk_baru.png"))
    entry_bg_6 = canvas.create_image(
        912.0,
        277.0,
        image=entry_image_6
    )
    tambah_produk_entry_6 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    tambah_produk_entry_6.place(
        x=883.0,
        y=250.0,
        width=58.0,
        height=52.0
    )

    canvas.create_text(
        868.0,
        206.0,
        anchor="nw",
        text="Ukuran",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1)
    )

    canvas.create_text(
        56.0,
        32.0,
        anchor="nw",
        text="Tambah Produk Baru",
        fill="#FFFFFF",
        font=("Segoe UI", 40 * -1, 'bold')
    )

    canvas.create_text(
        71.0,
        150.0,
        anchor="nw",
        text="Data Sepatu",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, 'bold')
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_produk_baru.png"))
    
    tambah_produk_button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), tambah_data_excel(file_excel,ent_tambah_data_excel(tambah_produk_entry_2,tambah_produk_entry_1,tambah_produk_entry_5,tambah_produk_entry_6, random.randint(10000000,99999999),tambah_produk_entry_4,tambah_produk_entry_3)),window.destroy(),menu()],
        bg="#01041A",
        activebackground="#01041A",
        relief="flat"
    )
    tambah_produk_button_1.place(
        x=115.0,
        y=554.0,
        width=841.0,
        height=56.0
    )
    window.resizable(False, False)
    window.mainloop()

def kasir():
    global harga
    global merek
    global seri
    global jumlah
    global ukuran
    window_kasir = Tk()

    window_kasir.geometry("1080x720")
    window_kasir.configure(bg = "#01041A")


    canvas = Canvas(
        window_kasir,
        bg = "#01041A",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_kasir.png"))
    image_1 = canvas.create_image(
        317.0,
        304.0,
        image=image_image_1
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2_kasir.png"))
    image_2 = canvas.create_image(
        816.0,
        247.0,
        image=image_image_2
    )

    image_image_3 = PhotoImage(
        file=relative_to_assets("image_3_kasir.png"))
    image_3 = canvas.create_image(
        842.0,
        452.0,
        image=image_image_3
    )

    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1_kasir.png"))
    entry_bg_1 = canvas.create_image(
        454.5,
        277.0,
        image=entry_image_1
    )
    kasir_entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    kasir_entry_1.place(
        x=375.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        360.0,
        206.0,
        anchor="nw",
        text="Merek",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets("entry_2_kasir.png"))
    entry_bg_2 = canvas.create_image(
        212.5,
        277.0,
        image=entry_image_2
    )
    kasir_entry_2 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    kasir_entry_2.place(
        x=133.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        206.0,
        anchor="nw",
        text="Jenis",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_3 = PhotoImage(
        file=relative_to_assets("entry_3_kasir.png"))
    entry_bg_3 = canvas.create_image(
        904.5,
        452.0,
        image=entry_image_3
    )
    kasir_entry_3 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    kasir_entry_3.place(
        x=847.0,
        y=425.0,
        width=115.0,
        height=52.0
    )

    canvas.create_text(
        701.0,
        437.0,
        anchor="nw",
        text="Jumlah :",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_4 = PhotoImage(
        file=relative_to_assets("entry_4_kasir.png"))
    entry_bg_4 = canvas.create_image(
        212.5,
        419.0,
        image=entry_image_4
    )
    kasir_entry_4 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    kasir_entry_4.place(
        x=133.0,
        y=392.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        348.0,
        anchor="nw",
        text="Seri",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5_kasir.png"))
    entry_bg_5 = canvas.create_image(
        417.5,
        419.0,
        image=entry_image_5
    )
    kasir_entry_5 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    kasir_entry_5.place(
        x=375.0,
        y=392.0,
        width=85.0,
        height=52.0
    )

    canvas.create_text(
        360.0,
        348.0,
        anchor="nw",
        text="Ukuran",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    canvas.create_text(
        56.0,
        32.0,
        anchor="nw",
        text="Kasir",
        fill="#FFFFFF",
        font=("Segoe UI", 40 * -1, 'bold')
    )

    canvas.create_text(
        71.0,
        150.0,
        anchor="nw",
        text="Pilih Sepatu",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, 'bold')
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_kasir.png"))
    kasir_button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), print(kasir_entry_2.get(), kasir_entry_1.get(),kasir_entry_4.get(),kasir_entry_5.get(),kasir_entry_3.get()), set_sepatu(kasir_entry_1.get(),kasir_entry_4.get(),kasir_entry_5.get(),kasir_entry_3.get()), bs.kurangi_sepatu(os.path.join(os.getcwd(), 'Persediaan1.xlsx'),pd.read_excel(os.path.join(os.getcwd(), 'Persediaan1.xlsx')),kasir_entry_2.get(),kasir_entry_1.get(),kasir_entry_4.get(),kasir_entry_5.get(),kasir_entry_3.get()),window_kasir.destroy(), metode_pembayaran()],
        bg="#01041A",
        activebackground="#01041A",
        relief="flat"
    )
    kasir_button_1.place(
        x=112.0,
        y=589.0,
        width=847.0,
        height=57.0
    )
    window_kasir.resizable(False, False)
    window_kasir.mainloop()
    



def load_and_display_excel_data():
    # Membuat jendela utama
    root = tk.Tk()
    root.title("Excel to GUI")

    # Membaca data dari file Excel menggunakan pandas
    file_path = os.path.join(os.getcwd(), 'Persediaan1.xlsx')
    df = pd.read_excel(file_path)

    # Membuat Treeview untuk menampilkan data Excel
    tree = ttk.Treeview(root)
    tree.pack(fill=tk.BOTH, expand=True)

    # Menambahkan kolom di Treeview
    tree["column"] = list(df.columns)
    tree["show"] = "headings"
    for col in tree["column"]:
        tree.heading(col, text=col)

    # Menambahkan baris data di Treeview
    for row in df.itertuples(index=False):
        tree.insert("", "end", values=row)

    # Menjalankan aplikasi
    root.mainloop()


def print_nota():
    
    global diskon
    global merek
    global seri
    global harga
    global jumlah
    global ukuran
    window_nota = Tk()

    window_nota.geometry("1080x720")
    window_nota.configure(bg = "#FFFFFF")


    canvas = Canvas(
        window_nota,
        bg = "#FFFFFF",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_nota.png"))
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), window_nota.destroy(), menu()],
        relief="flat"
    )
    button_1.place(
        x=163.0,
        y=621.0,
        width=753.0,
        height=62.0
    )

    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_nota.png"))
    image_1 = canvas.create_image(
        540.0,
        339.0,
        image=image_image_1
    )

    canvas.create_text(
        205.0,
        218.0,
        anchor="nw",
        text=":",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        205.0,
        268.0,
        anchor="nw",
        text=":",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        205.0,
        318.0,
        anchor="nw",
        text=":",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        205.0,
        368.0,
        anchor="nw",
        text=":",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        208.0,
        438.0,
        anchor="nw",
        text=":",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        74.0,
        318.0,
        anchor="nw",
        text="Harga satuan",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        74.0,
        438.0,
        anchor="nw",
        text="SUBTOTAL",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        74.0,
        368.0,
        anchor="nw",
        text="Diskon",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        350.0,
        70.0,
        anchor="nw",
        text="NOTA PEMBELIAN SEPATU",
        fill="#000000",
        font=("Segoe UI", 30 * -1, 'bold')
    )

    canvas.create_text(
        73.0,
        268.0,
        anchor="nw",
        text="Ukuran",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        76.0,
        218.0,
        anchor="nw",
        text="Produk ",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        225.0,
        218.0,
        anchor="nw",
        text=f"{merek} seri {seri} \t Jumlah \t : {jumlah} pasang",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        225.0,
        268.0,
        anchor="nw",
        text=f"{ukuran}",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        225.0,
        318.0,
        anchor="nw",
        text=f"{harga} \t Total :"+ str(int(harga)*int(jumlah)),
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        225.0,
        368.0,
        anchor="nw",
        text=f"{diskon}",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        225.0,
        438.0,
        anchor="nw",
        text= (float(harga) * float(jumlah)) - float(diskon),
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_rectangle(
        231.0,
        118.99999999999986,
        844.0,
        124.0,
        fill="#000000",
        outline="")

    canvas.create_rectangle(
        29.0,
        413.0,
        1045.00048828125,
        418.0,
        fill="#000000",
        outline="")
    window_nota.resizable(False, False)
    window_nota.mainloop()

def tradein():
    global diskon
    window_tradein = Tk()

    window_tradein.geometry("1080x720")
    window_tradein.configure(bg = "#01041A")


    canvas = Canvas(
        window_tradein,
        bg = "#01041A",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_tradein.png"))
    image_1 = canvas.create_image(
        317.0,
        343.0,
        image=image_image_1
    )

    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1_tradein.png"))
    entry_bg_1 = canvas.create_image(
        454.5,
        277.0,
        image=entry_image_1
    )
    entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    entry_1.place(
        x=375.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        360.0,
        206.0,
        anchor="nw",
        text="Merek",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets("entry_2_tradein.png"))
    entry_bg_2 = canvas.create_image(
        212.5,
        277.0,
        image=entry_image_2
    )
    entry_2 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    entry_2.place(
        x=133.0,
        y=250.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        206.0,
        anchor="nw",
        text="Jenis",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_3 = PhotoImage(
        file=relative_to_assets("entry_3_tradein.png"))
    entry_bg_3 = canvas.create_image(
        212.5,
        390.0,
        image=entry_image_3
    )
    entry_3 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    entry_3.place(
        x=133.0,
        y=363.0,
        width=159.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        319.0,
        anchor="nw",
        text="Seri",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_4 = PhotoImage(
        file=relative_to_assets("entry_4_tradein.png"))
    entry_bg_4 = canvas.create_image(
        333.5,
        495.0,
        image=entry_image_4
    )
    entry_4 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    entry_4.place(
        x=133.0,
        y=468.0,
        width=401.0,
        height=52.0
    )

    canvas.create_text(
        118.0,
        424.0,
        anchor="nw",
        text="Harga",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5_tradein.png"))
    entry_bg_5 = canvas.create_image(
        417.5,
        390.0,
        image=entry_image_5
    )
    entry_5 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        font=("Segoe UI", 25 * -1)
    )
    entry_5.place(
        x=375.0,
        y=363.0,
        width=85.0,
        height=52.0
    )

    canvas.create_text(
        360.0,
        319.0,
        anchor="nw",
        text="Ukuran",
        fill="#FFFFFF",
        font=("Segoe UI Regular", 25 * -1)
    )

    canvas.create_text(
        56.0,
        32.0,
        anchor="nw",
        text="Trade-In",
        fill="#FFFFFF",
        font=("Segoe UI", 40 * -1, 'bold')
    )

    canvas.create_text(
        71.0,
        150.0,
        anchor="nw",
        text="Data sepatu lama",
        fill="#FFFFFF",
        font=("Segoe UI", 25 * -1, 'bold')
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_tradein.png"))
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), print(entry_1.get(),entry_2.get(),entry_3.get(),entry_4.get(),entry_5.get()), set_diskon(entry_4.get()), window_tradein.destroy(), kasir()],
        relief="flat",
        bg = "#01041A",
        activebackground="#01041A"
    )
    button_1.place(
        x=112.0,
        y=589.0,
        width=847.0,
        height=57.0
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2_tradein.png"))
    image_2 = canvas.create_image(
        830.0,
        340.0,
        image=image_image_2
    )
    window_tradein.resizable(False, False)
    window_tradein.mainloop()

def metode_pembayaran():
    global merek
    global seri
    global ukuran
    global jumlah
    window_mp = Tk()

    window_mp.geometry("1080x720")
    window_mp.configure(bg = "#141414")


    canvas = Canvas(
        window_mp,
        bg = "#141414",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    canvas.create_text(
        72.99999999999994,
        133.0,
        anchor="nw",
        text="Pilih metode pembayaran",
        fill="#FFFFFF",
        font=("montserrat", 30 * -1, 'bold')
    )

    canvas.create_text(
        72.99999999999994,
        58.0,
        anchor="nw",
        text="Metode Pembayaran",
        fill="#0057FF",
        font=("montserrat", 50 * -1, 'bold')
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_metode.png"))
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"), window_mp.destroy(), qris()],
        relief="flat",
        bg="#141414",
        activebackground="#141414"
    )
    button_1.place(
        x=357.99999999999994,
        y=226.0,
        width=295.0,
        height=328.0
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2_metode.png"))
    button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_2 clicked"), window_mp.destroy(),print_nota()],
        relief="flat",
        bg="#141414",
        activebackground="#141414"
    )
    button_2.place(
        x=20.999999999999943,
        y=226.0,
        width=336.0,
        height=328.0
    )

    canvas.create_rectangle(
        745.0,
        0.0,
        1080.0,
        720.0,
        fill="#1F1F1F",
        outline="")

    canvas.create_text(
        786.0,
        80.0,
        anchor="nw",
        text="Pembelian",
        fill="#FFFFFF",
        font=("montserrat", 30 * -1, 'bold')
    )

    canvas.create_text(
        786.0,
        143.0,
        anchor="nw",
        text=f'{merek} seri {seri}',
        fill="#FFFFFF",
        font=("montserrat Regular", 20 * -1)
    )

    canvas.create_text(
        786.0,
        178.0,
        anchor="nw",
        text=f"Ukuran \t : {ukuran}",
        fill="#FFFFFF",
        font=("montserrat Regular", 20 * -1)
    )

    canvas.create_text(
        786.0,
        248.0,
        anchor="nw",
        text=f"Diskon \t : {diskon}",
        fill="#FFFFFF",
        font=("montserrat Regular", 20 * -1)
    )

    canvas.create_text(
        786.0,
        283.0,
        anchor="nw",
        text="Subtotal \t : "+str((float(harga) * float(jumlah)) - float(diskon)),
        fill="#FFFFFF",
        font=("montserrat", 20 * -1, 'bold')
    )

    canvas.create_text(
        784.0,
        213.0,
        anchor="nw",
        text=f"Jumlah \t : {jumlah}",
        fill="#FFFFFF",
        font=("montserrat Regular", 20 * -1)
    )
    window_mp.resizable(False, False)
    window_mp.mainloop()

def qris():
    window_qris = Tk()

    window_qris.geometry("1080x720")
    window_qris.configure(bg = "#141414")


    canvas = Canvas(
        window_qris,
        bg = "#141414",
        height = 720,
        width = 1080,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_qris.png"))
    image_1 = canvas.create_image(
        540.0,
        360.0,
        image=image_image_1
    )

    canvas.create_text(
        73.0,
        133.0,
        anchor="nw",
        text="Nontunai",
        fill="#000000",
        font=("Segoe UI", 30 * -1, 'bold')
    )

    canvas.create_text(
        73.0,
        208.0,
        anchor="nw",
        text="Trasnsfer bank ke",
        fill="#000000",
        font=("Segoe UI", 20 * -1, 'bold')
    )

    canvas.create_text(
        73.0,
        248.0,
        anchor="nw",
        text="Mandiri \n104976349648 \t \t a.n. Fadia Dian Ambarrizka",
        fill="#000000",
        font=("Segoe UI Regular", 20 * -1)
    )

    canvas.create_text(
        74.0,
        316.0,
        anchor="nw",
        text="BRI\n6632 0102 3871 530 \t a.n. Ferrel Rafi Elian",
        fill="#000000",
        font=("Segoe UI Regular", 20 * -1)
    )

    canvas.create_text(
        74.0,
        390.0,
        anchor="nw",
        text="BNI\n1786806019 \t \t a.n. Atha Nabilah Aurellia",
        fill="#000000",
        font=("Segoe UI Regular", 20 * -1)
    )

    canvas.create_text(
        71.0,
        58.0,
        anchor="nw",
        text="Metode Pembayaran",
        fill="#0057FF",
        font=("Segoe UI", 45 * -1, "bold")
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_qris.png"))
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [print("button_1 clicked"),window_qris.destroy(),print_nota()],
        relief="flat"
    )
    button_1.place(
        x=87.0,
        y=601.0,
        width=428.0,
        height=72.0
    )
    window_qris.resizable(False, False)
    window_qris.mainloop()

# Memanggil fungsi untuk memuat dan menampilkan data Excel

def set_sepatu(me, se, uk, ju):
    global merek
    global seri
    global ukuran
    global jumlah
    merek = me
    seri = se
    ukuran=uk
    jumlah=ju

def set_diskon(val):
    global diskon
    diskon = val
#==================================================================================================================================
#==================================================================================================================================
from openpyxl import load_workbook
from openpyxl import Workbook

def tambah_data_excel(file_path, data):
    try:
        # Coba membuka file yang sudah ada
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Jika file tidak ditemukan, buat file baru dengan sheet aktif
        workbook = Workbook()
        sheet = workbook.active
        # Menambahkan header jika file baru
        headers = ['jenis', 'merek', 'seri', 'ukuran', "kode",  'harga', 'jumlah']
        sheet.append(headers)

    # Menambahkan data baru ke dalam sheet
    sheet.append(data)

    # Menyimpan workbook
    workbook.save(file_path)
    print(f"Data berhasil ditambahkan ke {file_path}")

def ent_tambah_data_excel(tambahproduk_entry_1,tambahproduk_entry_2,tambahproduk_entry_3,tambahproduk_entry_4, kodeid,tambahproduk_entry_5,tambahproduk_entry_6):
    data_sepatu = [tambahproduk_entry_1.get(),tambahproduk_entry_2.get(),tambahproduk_entry_3.get(),tambahproduk_entry_4.get(), kodeid,tambahproduk_entry_5.get(),tambahproduk_entry_6.get()]
    return data_sepatu
# Contoh penggunaan
file_excel = os.path.join(os.getcwd(),'Persediaan1.xlsx')
#data_baru = ["Elektronik", "Samsung", "Galaxy S21", "6.2 inch", 10000000, 10]

#tambah_data_excel(file_excel, data_baru)

#==================================================================================================================================
import tkinter as tk
from tkinter import messagebox
import pandas as pd

def edit_data(file_name, kode_produk=None, merek=None, seri=None, ukuran=None, harga=None, jumlah=None):
    # Load data from Excel file
    df = pd.read_excel(file_name)
    
    # If kode_produk is provided, update merek, seri, and ukuran accordingly
    if kode_produk:
        kode_row = df[df['kode_produk'] == kode_produk].index
        df.loc[kode_row, ['merek', 'seri', 'ukuran']] = [merek, seri, ukuran]
    
    # If harga is provided, update harga
    if harga is not None:
        if kode_produk:
            df.loc[kode_row, 'harga'] = harga
        else:
            data_row = df[(df['merek'] == merek) & (df['seri'] == seri) & (df['ukuran'] == ukuran)].index
            df.loc[data_row, 'harga'] = harga
    
    # If jumlah is provided, update jumlah
    if jumlah is not None:
        if kode_produk:
            df.loc[kode_row, 'jumlah'] = jumlah
        else:
            data_row = df[(df['merek'] == merek) & (df['seri'] == seri) & (df['ukuran'] == ukuran)].index
            df.loc[data_row, 'jumlah'] = jumlah
    
    # Save the updated data to the Excel file
    df.to_excel(file_name, index=False)

harga = int(random.randint(75,200))*10000

def submit_edit(kode_entry,merek_entry,seri_entry,ukuran_entry,harga_entry,jumlah_entry):
    # Get user inputs
    kode_produk = kode_entry.get()
    merek = merek_entry.get()
    seri = seri_entry.get()
    ukuran = ukuran_entry.get()
    harga = harga_entry.get()
    jumlah = jumlah_entry.get()
    print(kode_produk,merek,seri, ukuran, harga, jumlah)
    # Call edit_data function with user inputs
    edit_data('Persediaan1.xlsx', kode_produk, merek, seri, ukuran, harga, jumlah)
    
    # Show confirmation message
    messagebox.showinfo("Info", "Data telah diubah.")

#==================================================================================================================================
'''
def beli_sepatu(data_sepatu, kasir_entry_2):
    jenis = kasir_entry_2.get()
    menu_list = tampilkan_menu_dari_data(data_sepatu, str(jenis))
    if menu_list:
        choice = int(input("Pilih sepatu yang ingin dibeli (masukkan nomor pilihan): "))
        if 1 <= choice <= len(menu_list):
            chosen_item = menu_list[choice - 1]
            while True:
                try:
                    quantity = int(input("Masukkan kuantitas: "))
                    if quantity > 0:
                        break
                    else:
                        print("Kuantitas harus lebih dari 0. Silakan coba lagi.")
                except ValueError:
                    print("Masukkan angka yang valid untuk kuantitas.")
            
            print("\nAnda memilih :")
            print(f"Nama Sepatu  : {chosen_item[1]} {chosen_item[2]}")
            print(f"Harga        : {chosen_item[5]}")
            print(f"Kuantitas    : {quantity}\n")
            
            pesanan.append({
                'jenis': jenis,
                'merek': chosen_item[1],
                'series': chosen_item[2],
                'ukuran': chosen_item[3],
                'harga': chosen_item[5],
                'kode': chosen_item[4],
                'kuantitas': quantity
            })
        else:
            print("Nomor pilihan tidak valid.")
    else:
        print("Jenis sepatu tidak ditemukan.")

def tampilkan_menu_dari_data(data_sepatu, jenis_menu):
    df_filtered = data_sepatu[data_sepatu['jenis'] == jenis_menu]
    menu_list = []
    count = 1
    for index, row in df_filtered.iterrows():
        jenis = row['jenis']
        merek = row['merek']
        series = row['series']
        ukuran = row['ukuran']
        kode = row['kode']
        harga = row['harga']
        jumlah = row['jumlah']
        print(f"Pilihan {count}:")
        print(f"  Merek: {merek}")
        print(f"  Series: {series}")
        print(f"  Ukuran: {ukuran}")
        print(f"  Harga: {harga}")
        print(f"  Jumlah: {jumlah}")
        print(f"  Kode: {kode}")
        print("=========================================")
        menu_list.append((jenis, merek, series, ukuran, kode, harga, jumlah))
        count += 1
    return menu_list

pesanan = []

'''

#===================================================================================================================================
import pandas as pd

def program_kasir(kasir_entry_2, kasir_entry_1, kasir_entry_4,kasir_entry_5, kasir_entry_3):
    # Membaca file Excel
    df = pd.read_excel(os.path.join(os.getcwd(), 'Persediaan1.xlsx'))

    # Meminta input dari pengguna
    jenis = kasir_entry_2.get()
    merek = kasir_entry_1.get()
    series = kasir_entry_4.get()
    ukuran = kasir_entry_5.get()
    jumlah = kasir_entry_3.get()

    # Mencari baris yang sesuai di DataFrame
    item = df[(df['jenis'] == jenis) &
              (df['merek'] == merek) &
              (df['series'] == series) &
              (df['ukuran'] == ukuran)]

    if not item.empty:
        harga = item['harga'].values[0]
        stok = item['jumlah'].values[0]

        if stok >= jumlah:
            total_harga = harga * jumlah
            sisa_stok = stok - jumlah
            df.loc[item.index, 'jumlah'] = sisa_stok
            df.to_excel('Persediaan1.xlsx', index=False)
            print(f"Harga total: {total_harga}")
            print(f"Sisa stok untuk {jenis} {merek} {series} ukuran {ukuran}: {sisa_stok}")
        else:
            print("Jumlah yang diminta melebihi stok yang tersedia.")
    else:
        print("Barang tidak ditemukan.")

import pandas as pd

def get_harga_by_series_and_size(series, ukuran):
    # Membaca data dari file Excel
    file_path=os.path.join(os.getcwd(), 'Persediaan1.xlsx')
    data_sepatu = pd.read_excel(file_path)

    # Melakukan pencarian berdasarkan series dan ukuran
    result = data_sepatu[(data_sepatu['series'] == series) & (data_sepatu['ukuran'] == ukuran)]

    if not result.empty:
        # Jika ada hasil, kembalikan harga
        harga = result.iloc[0]['harga']
        return harga
    
    else:
        return 2000000

# Contoh pemanggilan fungsi

# Memanggil fungsi kasir
merek = ''
seri = ''
ukuran = 0
jumlah = 0

diskon = 0
#==========================================================================================================================================
program_awal()#==================================================================================================================================



